import numpy as np
import matplotlib.pyplot as plt
import openpyxl as px


# Funcion para lectura y obtencion de datos de inicio
# La funcion devuelve la demanda de los clientes, capacidad de los vehiculos de primer nivel y segundo nivel, capacidad de los centros locales y regionales
def read_data(n_clientes, n_productos, n_periodos, n_vehiculos_p, n_vehiculos_s, n_centrosregionales, n_centroslocales):
    # total_columnas = variable que me permite moverme entre los indices de las hojas de datos
    total_columnas = n_productos*n_periodos

    # lectura y obtencion de datos de los clientes
    datos = px.load_workbook('datos.xlsx')                                          # carga de la hoja de excel de datos
    hoja_clientes = datos['clientes']                                               # seleccionar la hoja clientes como hoja activa
    # Obtencion de las demandas de la tabla de la hoja clientes segun la cantidad de clientes, productos y periodos
    demanda_clientes = [[hoja_clientes.cell(row=i, column=j).value for j in range(2, 2+total_columnas)] for i in range(3, 3+n_clientes)]

    # lectura y obtencion de datos de los vehiculos de primer y segundo nivel
    hoja_vehiculos = datos['vehiculos']                                             # seleccionar la hoja vehiculos como hoja activa
    # Obtencion de las demandas de los vehiculos de primer nivel de la primera tabla en la hoja vehiculos segun la cantidad de vehiculos de primer nivel y periodos
    capacidad_vehiculos_p = [[hoja_vehiculos.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(2, 2+n_vehiculos_p)]
    # Obtencion de las demandas de los vehiculos de segundo nivel de la segunda tabla en la hoja vehiculos segun la cantidad de periodos, vehiculos de primer y segundo nivel
    capacidad_vehiculos_s = [[hoja_vehiculos.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(3+n_vehiculos_p, 3+n_vehiculos_p+n_vehiculos_s)]

    # lectura y obtencion de datos de las instalaciones de primer y segundo nivel
    hoja_instalaciones = datos['instalaciones']                                     # seleccionar la hoja instalaciones como hoja activa
    # Obtencion de las capacidades de los centros de primer nivel de la primera tabla en la hoja instalaciones segun la cantidad de centros regionales y periodos
    capacidad_cr = [[hoja_instalaciones.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(3, 3+n_centrosregionales)]
    # Obtencion de las capacidades de los centros de segundo nivel de la segunda tabla en la hoja instalaciones segun la cantidad de centros locales, centros regionales y periodos
    capacidad_cl = [[hoja_instalaciones.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(5+n_centrosregionales, 5+n_centrosregionales+n_centroslocales)]

    # lectura y obtencion de datos del costo de mantener inventario para el primer escalon
    hoja_inventario = datos['inventario']
    # Obtencion de los costos en la hoja seleccionada segun a cantidad de centros regionales y productos
    costo_inventario = [[hoja_inventario.cell(row=i, column=j).value for j in range(2, 2+n_productos)] for i in range(2, 2+n_centrosregionales)]

    return np.array(demanda_clientes), np.array(capacidad_vehiculos_p), np.array(capacidad_vehiculos_s), np.array(capacidad_cr), np.array(capacidad_cl), np.array(costo_inventario)


# Funcion que binariza los valores que se pasen como parametro
# con el objetivo de llevar un control de la operacion de resta realizada en las capacidades de los centros y los vehiculos
def binarize(val):
    if val > 0:
        return 1
    else:
        return 0


# Funcion que traduce en un diccionario la matriz que se pasa como parametro
# con el objetivo de llevar un registro temporal de las asignaciones
def dictionarize(mat):
    dicc = {}                                                               # inicializacion de un diccionario vacio
    for idx, val in enumerate(mat[1, :]):                                   # recorrer cada uno de los valores de la matriz y enumerarlos para llevar un control de indices
        if val not in dicc.keys():
            dicc[val] = [mat[0, idx]]                                       # si el valor no esta en el diccionario se guarda y se le asigna el contenido correspondiente de la matriz
        else:
            dicc[val].append(mat[0, idx])                                   # en caso de que el valor si este en el diccionario se adjunta el contenido correspondiente de la matriz al ya asignado anteriormente
    return dicc


# Funcion que permite traducir o mapear las asignaciones de la matriz que se pasa como parametro
# con el objetivo de poder almacenar y utilizar los centros habilitados en cada asignacion
def maping(demanda):
    n_centros_habs = len(demanda)
    demanda_cl_np = np.array([x[1] for x in demanda])
    centros_habs = [int(x[0]) for x in demanda]
    return n_centros_habs, demanda_cl_np, centros_habs


# Funcion para las asignaciones (Decision de localizacion-asignacion) recibe como parametros:
# n_asignar: numero de clientes o centros locales que seran asignados segun el nivel
# n_centros: numero de centros locales o centros regionales a los que se asigna segun el nivel
# periodo: periodo en el que se encuentre la asignacion
# n_productos: numero de productos
# capacidad_centro: capacidad del centro al cual se le esten asignando clientes u otros centros segun el nivel
# demanda: demanda del cliente o centro segun el nivel
# mapeo: vector con los centros habilitados, para el segundo escalon en el periodo 1 la longitud del mapeo es igual a 0
# La funcion devuelve la matriz de localizacion-asignacion y la demanda del centro local o regional segun el nivel y el periodo que se este trabajando

def asignaciones(n_asignar, n_centros, periodo, n_productos, capacidad_centro, demanda, mapeo, escalon):

    por_asignar = np.array(range(1, n_asignar + 1))                             # creacion de una lista con los clientes o centros que se asignaran
    centros = np.array(range(1, n_centros + 1))                                 # creacion de una lista con los centros a los que se asginaran
    asignacion_lv = np.array([por_asignar, np.zeros(len(por_asignar))])         # inicializacion de la matriz localizacion-asignacion
    intentos = 0                                                                # variable para el control de intentos de asignaciones
    rango = (periodo-1)*n_productos                                             # rango de indices para moverse a traves de la matriz de demanda
    copia_capacidad = np.copy(capacidad_centro)                                 # copia de la capacidad del centro para evitar modificaciones en las capacidades originales
    if (len(mapeo) > 0 and escalon == 2) or (len(mapeo) > 0 and escalon == 1):
        centro_temp = np.random.choice(mapeo)                                   # selecciona un centro habilitado previamente en otro periodo
    else:
        centro_temp = np.random.choice(centros)                                 # seleccion aleatoria del primer centro - habilitacion del primer centro
    # idx_c = 0
    while len(por_asignar) > 0:                                                 # mientras existan clientes o centros por asignar
        if intentos < 3:                                                        # si los intentos de asignacion son menores a 3
            asig_temp = np.random.choice(por_asignar)                           # selecciona un cliente o centro aleatorio para asignar
            # resta la capacidad del centro con la demanda del cliente o centro
            if periodo > 1 and escalon == 1:
                resta_capacidad = copia_capacidad[centro_temp - 1, :] - demanda[asig_temp - 1, 0:n_productos]
                # actualiza la nueva capacidad del centro
                copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] - demanda[asig_temp - 1, 0:n_productos]
            else:
                resta_capacidad = copia_capacidad[centro_temp-1, :]-demanda[asig_temp-1, rango:rango+n_productos]
                # actualiza la nueva capacidad del centro
                copia_capacidad[centro_temp-1, :] = copia_capacidad[centro_temp-1, :]-demanda[asig_temp-1, rango:rango+n_productos]
            binvec = np.array([binarize(x) for x in resta_capacidad])           # se binariza la resta de la capacidad
            if binvec.all():                                                    # si la resta en la capacidad para todos los productos da un valor positivo
                idx_c = int(np.where(asignacion_lv[0, :] == asig_temp)[0])      # almacena el indice del centro
                asignacion_lv[1, idx_c] = centro_temp                           # guarda en la matriz de localizacion-asignacion el centro en la posicion del cliente o centro seleccionado
                idx_d = int(np.where(por_asignar == asig_temp)[0])              # almacena el indice del cliente o centro que ya fue asignado
                por_asignar = np.delete(por_asignar, [idx_d])                   # elimina de la lista el cliente o centro asignado
            else:                                                               # en caso de que la resta de la capacidad resulte negativa para al menos 1 valor
                intentos += 1                                                   # aumenta en 1 el numero de intentos
                # reestablece la capacidad del centro al momento antes de la resta
                if periodo > 1 and escalon == 1:
                    copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] + demanda[asig_temp - 1, 0:n_productos]
                else:
                    copia_capacidad[centro_temp - 1, :] = copia_capacidad[centro_temp - 1, :] + demanda[asig_temp - 1, rango:rango + n_productos]
        else:                                                                   # al llegar al numero maximo de intentos
            if (len(mapeo) > 1 and escalon == 2) or (len(mapeo) > 1 and escalon == 1):
                idx_cl = np.where(mapeo == centro_temp)
                mapeo = np.delete(mapeo, [idx_cl])
                idx_m = np.where(centros == centro_temp)
                centros = np.delete(centros, [idx_m])
                centro_temp = np.random.choice(mapeo)
                intentos = 0
            elif (len(mapeo) == 1 and escalon == 2) or (len(mapeo) == 1 and escalon == 1):                              # si se usaron todos los centros y aun hay clientes por asignar
                idx_cl = np.where(mapeo == centro_temp)                         # seleccionamos el indice del centro que ya agoto su capacidad
                mapeo = np.delete(mapeo, [idx_cl])
                centro_temp = np.random.choice(centros)                         # seleccionamos o habilitamos un nuevo centro que no se haya usado
                intentos = 0                                                    # reiniciamos el numero de intentos
            else:
                idx_cl = np.where(centros == centro_temp)                       # seleccionamos el indice del centro que ya agoto su capacidad
                centros = np.delete(centros, [idx_cl])                          # eliminamos de la lista de centros el centro que ya fue agotado
                centro_temp = np.random.choice(centros)                         # seleccionamos o habilitamos un nuevo centro
                intentos = 0                                                    # reiniciamos el numero de intentos

    dicc = dictionarize(asignacion_lv)                                          # generamos un diccionario con las asignaciones realizadas donde la llave es el centro y los valores con los centros o clientes asignados
    demandaf = []                                                               # inicializamos un vector donde se almacenaran la demandas finales de los centros
    for centro, asignados in dicc.items():                                      # para cada centro y valores asignados a ese centro
        demandacentro = [centro]                                                # creamos un vector con el centro seleccionado
        suma = 0                                                                # inicializamos la suma de las demandas
        for asig in asignados:
            if periodo > 1 and escalon == 1:
                suma += demanda[int(asig) - 1, 0:n_productos]
            else:
                suma += demanda[int(asig) - 1, rango:rango + n_productos]       # sumamos las demandas de cada cliente o centro que fue asignado a ese centro
        demandacentro.append(suma)                                              # adjuntamos la demanda al vector que contiene las demandas del centro
        demandaf.append(demandacentro)                                          # adjuntamos las demandas del centro al vector que contiene las demandas de todos los centros

    return asignacion_lv, demandaf


# Funcion para el plan de rutas, recibe como parametros:
# asignacion_lv: matriz de asignacion-localizacion
# n_vehiculos: numero de vehiculos
# demanda: demanda de clientes o centros segun corresponda el nivel
# periodo: periodo en el que se encuentre la asignacion de rutas
# n_productos: numero de productos
# La funcion devuelve una lista de listas con el plan de rutas del periodo que se este trabajando

def rutas(asignacion_lv, n_vehiculos, capacidad_vehiculos, demanda, periodo, n_productos, escalon):
    rutas_lv = []                                                         # inicializacion del vector de vetores de rutas
    vehiculos = list(range(1, n_vehiculos+1))                             # creacion de una lista de vehiculos con los vehiculos existentes
    dicci_asignacion = dictionarize(asignacion_lv)                        # generacion de un diccionario con la matriz de asignacion-localizacion
    rango = (periodo-1)*n_productos                                       # rango de indices para moverse a travesde la matriz de demandas
    capacidad_vehiculos_copy = np.copy(capacidad_vehiculos)
    for centro, asignados in dicci_asignacion.items():                    # por cada centro y sus respectivas asignaciones
        idx_c = 0                                                         # inicializacion del indice para recorrer los centros o clientes asignados
        vehiculo_temp = np.random.choice(vehiculos)                       # seleccionamos un vehiculo aleatorio
        ruta_temp = [int(centro), vehiculo_temp, 0]                       # ingresamos el centro, el vehiculo e iniciamos ruta
        veh_cap = np.copy(capacidad_vehiculos_copy[vehiculo_temp-1, :])        # copiamos la capacidad del vehiculo para evitar modificar la capacidad orginal
        while idx_c < len(asignados):                                     # mientras no se hayan recorrido todos los asignados(cliente o centro segun corresponda)
            if periodo > 1 and escalon == 1:
                dem_c = demanda[idx_c, 0:n_productos]                     # obtenemos la demanda del asignado
            else:
                dem_c = demanda[idx_c, rango:rango+n_productos]           # obtenemos la demanda del asignado
            resta = veh_cap - dem_c                                       # restamos la capacidad del vehiculo con la demanda del asignado
            binvec = np.array([binarize(x) for x in resta])               # binarizamos la resta
            if binvec.all():                                              # si la resta es positiva para cada producto
                ruta_temp.append(asignados[idx_c])                        # agregue a la ruta el centro accediendo al indice del mapeo que le corresponde
                veh_cap -= dem_c                                          # restamos la capacidad del vehiculo
                idx_c += 1                                                # aumentamos el indice
            else:                                                         # al agotar la capacidad del vehiculo
                if ruta_temp[-1] == 0:                                    # si el vehiculo seleccionado no satisface ningun cliente
                    ruta_temp.pop(-1)                                     # elimina el 0 de inicio de ruta de ese vehiculo
                    ruta_temp.pop(-1)                                     # elimina el vehiculo
                else:
                    ruta_temp.append(0)                                   # finalizamos ruta
                    vehiculos.pop(vehiculos.index(vehiculo_temp))         # eliminamos el vehiculo asigando de la lista de vehiculos
                    vehiculo_temp = np.random.choice(vehiculos)           # seleccionamos un nuevo vehiculo aleatorio
                    veh_cap = capacidad_vehiculos_copy[vehiculo_temp - 1, :]   # obtenemos la capacidad del nuevo vehiculo
                    ruta_temp += [vehiculo_temp, 0]                       # agregamos el vehiculo al plan de rutas e iniciamos una nuvea ruta
        vehiculos.pop(vehiculos.index(vehiculo_temp))
        ruta_temp.append(0)                                               # finalizamos ruta
        rutas_lv.append(ruta_temp)                                        # agregamos la ruta completa a la lista de rutas

    return rutas_lv

# Funcion para la generacion de un individuo completo
# se recibe como parametros
# n_clientes : numero de clientes
# n_centros_locales: numero de centros locales
# n_centros_regionales: numero de centros regionales
# n_periodos: numero de periodos
# n_productos: nuymero de productos
# n_vehiculos_s: numero de vehiculos de segundo escalon
# n_vehiculos_p: numero de vehiculos de primer escalon
# capacidad_cl: matriz con las capacidades de los centros locales por producto
# capacidad_cr: matriz con las capacidades de los centros regionales por producto
# capacidad_vehiculos_p: matriz de capacidad de carga de los vehiculos de primer escalon por producto
# capacidad_vehiculos_s: matriz de capacidad de carga de los vehiculos de segundo escalon por producto
# demanda_cl: matriz de demanda de los clientes por producto y periodo
# La funcion devuelve o retorna:
# asignaciones_primer_lv: matriz de asignaciones de primer escalon
# asignaciones_segundo_lv: matriz de asignaciones de segundo escalon
# rutas_primer_lv: lista de listas con las rutas de primer escalon
# rutas_segundo_lv: lista de listas con las rutas de segundo escalon
# demanda_cr_full: lista de diccionarios con las demandas de cada centro regional por producto, cada elemento de la lista corresponde a un periodo


def individuo(n_clientes, n_centroslocales, n_centrosregionales, n_periodos, n_productos, n_vehiculos_s, n_vehiculos_p, capacidad_cl, capacidad_cr, capacidad_vehiculos_p, capacidad_vehiculos_s, demanda_clientes):

    asignaciones_segundo_lv = []
    rutas_segundo_lv = []
    asignaciones_primer_lv = []
    rutas_primer_lv = []
    cl_habs = []
    cr_habs = []
    demanda_cr_full = []

    for perioactual in range(1, n_periodos + 1):
        dicti = {}
        # llamada a la funcion de asignacion para la asignacion-localizacion del segundo escalon
        asignacion_segundo_nivel, demanda_cl = asignaciones(n_clientes, n_centroslocales, perioactual, n_productos, capacidad_cl, demanda_clientes, cl_habs, 2)
        # variables de mapeo de los centros locales habilitados en el segundo escalon
        n_cl_habs, demanda_cl_np, cl_habs = maping(demanda_cl)
        # llamada a la funcion de rutas para la creacion del plan de rutas del segundo escalon
        rutas_segundo_nivel = rutas(asignacion_segundo_nivel, n_vehiculos_s, capacidad_vehiculos_s, demanda_clientes, perioactual, n_productos, 2)
        # llamada a la funcion de asignacion para la asignacion-localizacion del primer escalon
        asignacion_primer_nivel, demanda_cr = asignaciones(n_cl_habs, n_centrosregionales, perioactual, n_productos, capacidad_cr, demanda_cl_np, cr_habs, 1)
        asignacion_primer_nivel[0, :] = cl_habs
        # variables de mapeo de los centros locales habilitados en el primer escalon
        n_cr_habs, demanda_cr_np, cr_habs = maping(demanda_cr)
        # llamada a la funcion de rutas para la creacion del plan de rutas del primer escalon
        rutas_primer_nivel = rutas(asignacion_primer_nivel, n_vehiculos_p, capacidad_vehiculos_p, demanda_cl_np, perioactual, n_productos, 1)
        asignaciones_segundo_lv.append(asignacion_segundo_nivel)
        rutas_segundo_lv = rutas_segundo_lv + rutas_segundo_nivel
        asignaciones_primer_lv.append(asignacion_primer_nivel)
        rutas_primer_lv = rutas_primer_lv + rutas_primer_nivel
        for cr in demanda_cr:
            if cr[0] not in dicti.keys():
                dicti[cr[0]] = cr[1]
        demanda_cr_full.append(dicti)

    return asignaciones_primer_lv, asignaciones_segundo_lv, rutas_primer_lv, rutas_segundo_lv, demanda_cr_full


def fun_inventario(demandas_cr_full, n_periodos, n_productos, n_centrosregionales, capacidad_cr, inventario):

    matriz_demanda = []
    for p in range(n_periodos):
        dic_i = demandas_cr_full[p]
        matriz_demanda.append([])
        matriz_demanda[p] = [np.zeros([1, n_productos]) for _ in range(n_centrosregionales)]
        matriz_demanda[p] = [x[0] for x in matriz_demanda[p]]
        for j, k in dic_i.items():
            matriz_demanda[p][int(j - 1)] = np.array(k)
    matriz_demanda = np.array(matriz_demanda)
    centroshabs = []
    for c in range(n_centrosregionales):
        if np.sum(matriz_demanda[:, c, :]):
            centroshabs.append(c)
    centroshabs = np.array(centroshabs)
    matriz_demanda = matriz_demanda[:, centroshabs, :]
    valoresQ = {}
    valoresI = {}
    # para cada centro regional
    for u in range(len(centroshabs)):
        # para cada producto
        valoresQ[centroshabs[u] + 1] = []
        valoresI[centroshabs[u] + 1] = []
        for l in range(n_productos):
            # lista de productos por periodo del centro
            pr_per = matriz_demanda[:, u, l]
            carga0 = np.sum(pr_per)
            cap_centro = capacidad_cr[centroshabs[u], l]
            for w in range(n_periodos):
                d_cl = pr_per[w]
                if w == 0:
                    carga = carga0
                    iper1 = inventario[centroshabs[u], l]
                    if (iper1 >= d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(0, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(0, carga - iper1)
                    elif (iper1 < d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(d_cl - iper1, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(d_cl - iper1, carga - iper1)
                    elif (iper1 >= d_cl) and (carga <= cap_centro):
                        Q = np.random.randint(0, carga - iper1)
                    elif (iper1 < d_cl) and (carga <= cap_centro):
                        if d_cl - iper1 == carga - iper1:
                            Q = d_cl - iper1
                        else:
                            Q = np.random.randint(d_cl - iper1, carga - iper1)
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)
                    iper1 = It
                    carga = carga - d_cl - iper1
                elif 1 <= w < n_periodos - 1:
                    if w != 1:
                        carga = carga - Q
                    if (iper1 >= d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl <= carga:
                            Q = np.random.randint(0, cap_centro + d_cl - iper1)
                        else:
                            Q = np.random.randint(0, carga)
                    elif (iper1 < d_cl) and (carga > cap_centro):
                        if cap_centro + d_cl - iper1 <= carga:
                            Q = np.random.randint(d_cl - iper1, cap_centro + d_cl)
                        else:
                            Q = np.random.randint(d_cl - iper1, carga)
                    elif (iper1 >= d_cl) and (carga <= cap_centro):
                        if carga == 0:
                            Q = carga
                        else:
                            Q = np.random.randint(0, carga)
                    elif (iper1 < d_cl) and (carga <= cap_centro):
                        if d_cl - iper1 == carga:
                            Q = carga
                        else:
                            Q = np.random.randint(d_cl - iper1, carga)
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)
                    iper1 = It
                elif w == n_periodos - 1:
                    carga = carga - Q
                    Q = carga
                    valoresQ[centroshabs[u] + 1].append(Q)
                    It = iper1 + Q - d_cl
                    valoresI[centroshabs[u] + 1].append(It)

    return valoresQ, valoresI
